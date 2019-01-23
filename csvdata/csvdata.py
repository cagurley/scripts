# -*- coding: utf-8 -*-
"""
Created on Fri Jan 18 15:29:37 2019

@author: cagurl01
"""

import csv
import re
from collections import OrderedDict
from datetime import datetime
from openpyxl import load_workbook

# NOTE: May need to include openpyxl module import above
# Cell merge patch start
from openpyxl.worksheet import Worksheet
from openpyxl.reader.worksheet import WorkSheetParser
from openpyxl.worksheet.merge import MergeCells
from openpyxl.worksheet.cell_range import CellRange
# from openpyxl.utils import range_boundaries ### Included in monkey patch but unused


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    """
    # apply patch 1
    def merge_cells(self, range_string=None, start_row=None,
                    start_column=None, end_row=None, end_column=None):
        cr = CellRange(range_string=range_string, min_col=start_column,
                       min_row=start_row, max_col=end_column, max_row=end_row)
        self.merged_cells.add(cr.coord)
        # self._clean_merge_range(cr)

    Worksheet.merge_cells = merge_cells

    # apply patch 2
    def parse_merge(self, element):
        merged = MergeCells.from_tree(element)
        self.ws.merged_cells.ranges = merged.mergeCell
        # for cr in merged.mergeCell:
        #     self.ws._clean_merge_range(cr)

    WorkSheetParser.parse_merge = parse_merge


patch_worksheet()
# Patch end


def date_sub(date_values, string):
    try:
        if (not hasattr(date_values, '__iter__')
                or len(date_values) != 3
                or not re.match(r'\d{2}', date_values[0])
                or not re.match(r'\d{2}', date_values[1])
                or not re.match(r'\d{4}', date_values[2])):
            raise FormatError("Must provide an iterable of three digit-only text values in the form of ('MM', 'DD', 'YYYY').")
        if not isinstance(string, str):
            raise TypeError('Must provide string for substitution.')
    except (FormatError, TypeError) as e:
        print(str(e))
        raise e
    else:
        return (string.replace('MM', date_values[0])
                .replace('DD', date_values[1])
                .replace('YYYY', date_values[2]))


def dupe_compare_raw(set_with_raw, primary_key, other_with_raw, other_key):
    primary_values = []
    other_values = []
    dupes = set()
    for row in set_with_raw.raw_data:
        primary_values.append(row[primary_key])
    for row in other_with_raw.raw_data:
        other_values.append(row[other_key])
    for pvalue in primary_values:
        if pvalue in other_values:
            dupes.add(pvalue)
    return dupes


def split_datestring(datestring):
    try:
        datestring = str(datestring.lower())
        if datestring == 'now':
            return datetime.today().strftime('%m %d %Y').split()
        if re.match(r'\d{2}/\d{2}/\d{4}', datestring):
            return datestring.split('/')
        else:
            raise FormatError("Enter date as 'MM/DD/YYYY' or the word 'now'.")
    except FormatError as e:
        print(str(e))
        raise e


class FormatError(Exception):
    def __init__(self, message):
        super().__init__(message)


class CSVDataset:
    def __init__(self, filepath):
        self.filepath = filepath
        self.raw_data = []
        self.sum_count = []

    def calc_union_column(self, cond_dict, header='',
                          in_union='', not_in_union=''):
        for (index, row) in enumerate(self.raw_data):
            match = False
            for (key, value) in cond_dict.items():
                if row[key] == value:
                    self.raw_data[index].update({header: in_union})
                    match = True
                    break
            if not match:
                self.raw_data[index].update({header: not_in_union})

    def csv_to_raw(self):
        with open(self.filepath) as data:
            dreader = csv.DictReader(data)
            for row in dreader:
                self.raw_data.append(row)
        return None

    def data_to_csv(self, filepath, sum_count=False):
        with open(filepath, 'w', newline='') as output:
            if sum_count is True:
                dwriter = csv.DictWriter(output, self.sum_count[0].keys())
                dwriter.writeheader()
                dwriter.writerows(self.sum_count)
            else:
                dwriter = csv.DictWriter(output, self.raw_data[0].keys())
                dwriter.writeheader()
                dwriter.writerows(self.raw_data)
        return None
    
    
    def dupe_scan_raw(self, primary_key):
        primary_values = []
        dupes = set()
        for row in self.raw_data:
            primary_values.append(row[primary_key])
        for value in primary_values:
            if primary_values.count(value) > 1:
                dupes.add(value)
        return dupes
    

    def lookup_from_minor_raw(self, xlsx_minor_set, self_key, match_key,
                              header='', found='', not_found='', return_key=None):
        for (sindex, srow) in enumerate(self.raw_data):
            end = False
            for mrow in xlsx_minor_set.raw_data:
                if srow[self_key] == mrow[match_key]:
                    if return_key is not None:
                        self.raw_data[sindex].update({header: mrow[return_key]})
                    else:
                        self.raw_data[sindex].update({header: found})
                    end = True
                    break
            if not end:
                self.raw_data[sindex].update({header: not_found})
        return None

    def raw_to_sum(self, key=None, filter_key=None, filter_value=None):
        unique_values = set()
        for row in self.raw_data:
            value = row[key]
            if value.isdigit():
                value = int(value)
            unique_values.add(value)
        unique_values = list(unique_values)
        unique_values.sort()

        for (index, value) in enumerate(unique_values):
            unique_values[index] = {
                key: value,
                'COUNT': 0
            }
        if filter_key is not None and filter_value is not None:
            raw = []
            for row in self.raw_data:
                if row[filter_key] == filter_value:
                    raw.append(row)
        else:
            raw = self.raw_data
        for body_row in raw:
            for unique_row in unique_values:
                if body_row[key] == str(unique_row[key]):
                    unique_row['COUNT'] += 1
        self.sum_count = unique_values
        return None


class XLSXMinorSet:
    def __init__(self, filepath):
        self.filepath = filepath
        self.raw_data = []

    def xlsx_to_raw(self):
        wb = load_workbook(self.filepath)
        col_counter = 0
        rows = []
        for scol in wb.active.iter_cols():
            end = False
            for ccell in scol:
                if ccell.value:
                    col_counter += 1
                    break
                end = True
            if end:
                break
        for srow in wb.active.iter_rows(max_col=col_counter):
            end = False
            for rcell in srow:
                if rcell.value:
                    rows.append(srow)
                    break
                end = True
            if end:
                break
        header = rows.pop(0)
        for body_row in rows:
            row_dict = OrderedDict()
            for (index, hcell) in enumerate(header):
                row_dict.update({hcell.value: body_row[index].value})
            self.raw_data.append(row_dict)
        return None
