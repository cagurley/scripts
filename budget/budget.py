# -*- coding: utf-8 -*-
"""
Created on Mon Jan 14 16:35:37 2019

@author: cagurl01
"""

import importlib.machinery as impmach
from importlib.util import module_from_spec
loader = impmach.SourceFileLoader('csvdata', '../csvdata/csvdata.py')
spec = impmach.ModuleSpec('csvdata', loader)
csvdata = module_from_spec(spec)
loader.exec_module(csvdata)

import openpyxl as opx
import os
from copy import copy
from configparser import ConfigParser

# NOTE: May need to include openpyxl module import above
# Cell merge patch start
from openpyxl.worksheet import Worksheet
from openpyxl.reader.worksheet import WorkSheetParser
from openpyxl.worksheet.merge import MergeCells
from openpyxl.worksheet.cell_range import CellRange
# from openpyxl.utils import range_boundaries ### Included in monkey patch but unused


class AccessError(Exception):
    def __init__(self, message):
        super().__init__(message)


class ContentError(Exception):
    def __init__(self, message):
        super().__init__(message)


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    """
    # apply patch 1
    def merge_cells(self, range_string=None, start_row=None,
                    start_column=None, end_row=None, end_column=None):
        cr = CellRange(range_string=range_string, min_col=start_column,
                       min_row=start_row, max_col=end_column, max_row=end_row)
        self.merged_cells.add(cr.coord)
        # self._clean_merge_range(cr)

    Worksheet.merge_cells = merge_cells

    # apply patch 2
    def parse_merge(self, element):
        merged = MergeCells.from_tree(element)
        self.ws.merged_cells.ranges = merged.mergeCell
        # for cr in merged.mergeCell:
        #     self.ws._clean_merge_range(cr)

    WorkSheetParser.parse_merge = parse_merge


patch_worksheet()
# Patch end


def extract_from_ini(filepath, sect_opt_dict):
    try:
        parser = ConfigParser()
        parser.read(filepath)
        for section in sect_opt_dict.keys():
            if not parser.has_section(section):
                raise ContentError("Required section '[{}]' not found in '{}'.".format(section, filepath))
            for option in sect_opt_dict[section]:
                if not parser.has_option(section, option):
                    raise ContentError("Required option '{}' not found in section [{}] of '{}'.".format(option, section, filepath))
    except ContentError as e:
        print(str(e))
        raise e
    else:
        ini_values = []
        for section in sect_opt_dict.keys():
            sect_values = []
            for option in sect_opt_dict[section]:
                sect_values.append(parser.get(section, option))
            ini_values.append(tuple(sect_values))
        return ini_values


def validate_file(filepath, mode, extension=None):
    try:
        if isinstance(os.altsep, str):
            filepath = filepath.replace(os.altsep, os.sep)
            dirpath = os.path.split(filepath)[0]
            fileext = os.path.splitext(filepath)[1]
        if dirpath and not os.path.exists(dirpath):
            raise AccessError("'{}': Path does not exist.".format(dirpath))
        elif fileext != None and fileext != extension:
            raise AccessError("'{}': Extension does not match '{}'.".format(filepath, extension))
        else:
            if mode != os.F_OK and not os.access(filepath, mode):
                raise AccessError("'{}': Cannot access file in appropriate mode. Check directory and file permissions.".format(filepath))
            elif mode == os.F_OK and os.access(filepath, mode):
                overwrite = input("'{}': File already exists. Continue with overwrite? [y/N] ".format(filepath))
                if overwrite.lower() not in ('y', 'yes'):
                    raise AccessError('Overwrite aborted.')
                    return None
    except AccessError as e:
        print(str(e))
        raise e
    else:
        return None


INI_LAYOUT = {'FILEPATHS': ('PORTERDATA', 'ADSCHADATA', 'HOUSINGDATA',
                            'BUDGETTEMPLATE', 'BUDGETOUTPUT')}
pfp, afp, hfp, btfp, bofp = ('', '', '', '', '')
date_values = None
menu = True
while menu:
    try:
        validate_file(r'budget.ini', os.R_OK, '.ini')
        print("...Reading from 'budget.ini'...")
        extracted = extract_from_ini(r'budget.ini', INI_LAYOUT)
    except (AccessError, ContentError):
        print("""\nYou must have a file named 'budget.ini' in the current
working directory with the layout below between stars.
Replace each 'path' with the appropriate unquoted filepath:

********************

[FILEPATHS]
PORTERDATA=path
ADSCHADATA=path
HOUSINGDATA=path
BUDGETTEMPLATE=path
BUDGETOUTPUT=path

********************\n""")
        retry = input('Retry? [Y/n] ')
        if retry.lower() in ('n', 'no'):
            menu = False
            print('Operation aborted.')
            break
    else:
        pfp, afp, hfp, btfp, bofp = extracted[0]
        break
while menu:
    try:
        other_date = input('Should this report be for a date other than today? [y/N] ')
        if other_date.lower() in ('y', 'yes'):
            input_date = input("Enter report date as 'MM/DD/YYYY': ")
            date_values = csvdata.split_datestring(input_date)
        else:
            date_values = csvdata.split_datestring('now')
        bofp = csvdata.date_sub(date_values, bofp)
    except csvdata.FormatError:
        pass
    else:
        break
while menu:
    try:
        validate_file(pfp, os.R_OK, '.csv')
        validate_file(afp, os.R_OK, '.csv')
        validate_file(hfp, os.R_OK, '.xlsx')
        validate_file(btfp, os.R_OK, '.xlsx')
        validate_file(bofp, os.F_OK, '.xlsx')
    except AccessError:
        retry = input('Retry? [Y/n] ')
        if retry.lower() in ('n', 'no'):
            menu = False
            print('Operation aborted.')
            break
    else:
        pset = csvdata.CSVDataset(pfp)
        aset = csvdata.CSVDataset(afp)
        hset = csvdata.XLSXMinorSet(hfp)
        print('\n...Extracting data...')
        pset.csv_to_raw()
        aset.csv_to_raw()
        hset.xlsx_to_raw()
        print('\n...Checking for duplicates...')
        pdupes = pset.dupe_scan_raw('EMPLID')
        if pdupes:
            print('Found duplicate emplids in Porter data:\n')
            for value in pdupes:
                print('\t{}'.format(value))
            cont = input('\nContinue? [y/N] ')
            if cont.lower() not in ('y', 'yes'):
                menu = False
                print('Operation aborted.')
                break
        adupes = aset.dupe_scan_raw('EMPLID')
        if adupes:
            print('Found duplicate emplids in ADSCHA data:\n')
            for value in adupes:
                print('\t{}'.format(value))
            cont = input('\nContinue? [y/N] ')
            if cont.lower() not in ('y', 'yes'):
                menu = False
                print('Operation aborted.')
                break
        xdupes = csvdata.dupe_compare_raw(pset, 'EMPLID', aset, 'EMPLID')
        if xdupes:
            print('Found emplids shared between Porter and ADSCHA datasets:\n')
            for value in xdupes:
                print('\t{}'.format(value))
            cont = input('\nContinue? [y/N] ')
            if cont.lower() not in ('y', 'yes'):
                menu = False
                print('Operation aborted.')
                break
        print('...Appending columns...')
        pset.lookup_from_minor_raw(hset, 'EMPLID', 'ID Number', 'HOUS_APPL',
                                   'Has Application', 'No Application')
        aset.lookup_from_minor_raw(hset, 'EMPLID', 'ID Number', 'HOUS_APPL',
                                   'Has Application', 'No Application')
        UNION_COL_PARAMS = [
            [
                {
                    'HOUS_ASSIGN': 'Has Assignment',
                    'HOUS_APPL': 'Has Application',
                    'AFYLE': 'AFYLE'
                },
                'HOUS_INDICATED',
                'Housing Indicated',
                'No Indication'
            ],
            [
                {
                    'FRDP': 'FRDP',
                    'HOUS_INDICATED': 'Housing Indicated'
                },
                'HOUS/FRDP_INDICATED',
                'HOUS/FRDP Indicated',
                'No Indication'
            ],
            [
                {
                    'AWARD_STATUS': 'Accepted',
                    'HOUS/FRDP_INDICATED': 'HOUS/FRDP Indicated'
                },
                'ACCEPTED/HOUS/FRDP',
                'Accepted/HOU/FRDP',
                'No Indication'
            ]
        ]
        for param_set in UNION_COL_PARAMS:
            pset.calc_union_column(*param_set)
            aset.calc_union_column(*param_set)
#        pset.data_to_csv(r'test_appended_PORTER.csv')
#        aset.data_to_csv(r'test_appended_ADSCHA.csv')
        
        print('...Opening template...')
        report = opx.load_workbook(btfp)
        bottom_borders = []
        for bottom_cell in report['Detail'][82]:
            bottom_borders.append(copy(bottom_cell.border))
        print('...Writing header...')
        header = csvdata.date_sub(date_values, report['Detail']['B1'].value)
        report['Detail']['B1'].value = header
        print('...Calculating and writing body...')
        pset.raw_to_sum('AWARD_GROUP')
        aset.raw_to_sum('ADSCHA_AMT')
        for gcell in report['Detail']['B']:
            for group in pset.sum_count:
                if gcell.value == group['AWARD_GROUP']:
                    report['Detail'].cell(gcell.row, gcell.col_idx + 4, group['COUNT'])
        for group in aset.sum_count:
            for gcell in report['Detail']['B']:
                if gcell.value == 'Competitive Sum' and not report['Detail'].cell(gcell.row, gcell.col_idx + 2).value:
                    report['Detail'].cell(gcell.row, gcell.col_idx + 2, group['ADSCHA_AMT'])
                    report['Detail'].cell(gcell.row, gcell.col_idx + 4, group['COUNT'])
                    break
        SUM_PARAMS = [
            (
                ('AWARD_GROUP', 'AWARD_STATUS', 'Accepted'),
                ('ADSCHA_AMT', 'AWARD_STATUS', 'Accepted'),
                8
            ),
            (
                ('AWARD_GROUP', 'ACCEPTED/HOUS/FRDP', 'Accepted/HOU/FRDP'),
                ('ADSCHA_AMT', 'ACCEPTED/HOUS/FRDP', 'Accepted/HOU/FRDP'),
                10
            ),
            (
                ('AWARD_GROUP', 'STDNT_ENRL_STATUS', 'E'),
                ('ADSCHA_AMT', 'STDNT_ENRL_STATUS', 'E'),
                12
            ),
        ]
        for param_tuple in SUM_PARAMS:
            pset.raw_to_sum(*param_tuple[0])
            aset.raw_to_sum(*param_tuple[1])
            for gcell in report['Detail']['B']:
                for group in pset.sum_count:
                    if gcell.value == group['AWARD_GROUP']:
                        report['Detail'].cell(gcell.row, gcell.col_idx + param_tuple[2], group['COUNT'])
            for gcell in report['Detail']['B']:
                for group in aset.sum_count:
                    if (gcell.value == 'Competitive Sum'
                            and report['Detail'].cell(gcell.row, gcell.col_idx + 2).value == group['ADSCHA_AMT']):
                        report['Detail'].cell(gcell.row, gcell.col_idx + param_tuple[2], group['COUNT'])
                        break
        print('...Writing and formatting footer...')
        for gcell in report['Detail']['B']:
            if (gcell.value == 'Competitive Sum'
                    and not report['Detail'].cell(gcell.row, gcell.col_idx + 2).value):
                report['Detail'].delete_rows(gcell.row, 82 - gcell.row + 1)
                report['Detail'].delete_rows(gcell.row + 9, report['Detail'].max_row)
                for (index, bottom_cell) in enumerate(report['Detail'][gcell.row - 1]):
                    bottom_cell.border = bottom_borders[index]
                ranges_to_merge = [
                    'B{}:E{}'.format(gcell.row + 1, gcell.row + 2),
                    'B{0}:E{0}'.format(gcell.row + 3),
                    'B{0}:E{0}'.format(gcell.row + 4),
                    'B{0}:E{0}'.format(gcell.row + 5),
                    'B{0}:E{0}'.format(gcell.row + 6),
                    'B{0}:E{0}'.format(gcell.row + 7),
                    'F{0}:I{0}'.format(gcell.row + 1),
                    'J{0}:M{0}'.format(gcell.row + 1),
                    'F{0}:G{0}'.format(gcell.row + 2),
                    'F{0}:G{0}'.format(gcell.row + 3),
                    'F{0}:G{0}'.format(gcell.row + 4),
                    'F{0}:G{0}'.format(gcell.row + 5),
                    'F{0}:G{0}'.format(gcell.row + 6),
                    'F{0}:G{0}'.format(gcell.row + 7),
                    'H{0}:I{0}'.format(gcell.row + 2),
                    'H{0}:I{0}'.format(gcell.row + 3),
                    'H{0}:I{0}'.format(gcell.row + 4),
                    'H{0}:I{0}'.format(gcell.row + 5),
                    'H{0}:I{0}'.format(gcell.row + 6),
                    'H{0}:I{0}'.format(gcell.row + 7),
                    'J{0}:K{0}'.format(gcell.row + 2),
                    'J{0}:K{0}'.format(gcell.row + 3),
                    'J{0}:K{0}'.format(gcell.row + 4),
                    'J{0}:K{0}'.format(gcell.row + 5),
                    'J{0}:K{0}'.format(gcell.row + 6),
                    'J{0}:K{0}'.format(gcell.row + 7),
                    'L{0}:M{0}'.format(gcell.row + 2),
                    'L{0}:M{0}'.format(gcell.row + 3),
                    'L{0}:M{0}'.format(gcell.row + 4),
                    'L{0}:M{0}'.format(gcell.row + 5),
                    'L{0}:M{0}'.format(gcell.row + 6),
                    'L{0}:M{0}'.format(gcell.row + 7)
                ]
                for mrange in ranges_to_merge:
                    report['Detail'].merge_cells(mrange)
                for sum_cell in report['Detail']['F22:O22'][0]:
                    value = '=SUM({0}23:{0}{1})'.format(sum_cell.column, gcell.row)
                    report['Detail'].cell(sum_cell.row, sum_cell.col_idx, value)
                for sum_cell in report['Detail']['H{}:H{}'.format(gcell.row + 4, gcell.row + 7)]:
                    sum_cell = sum_cell[0]
                    report['Detail']['H{}'.format(sum_cell.row)].value = '=$F${}-F{}'.format(gcell.row + 3, sum_cell.row)
                for sum_cell in report['Detail']['L{}:L{}'.format(gcell.row + 4, gcell.row + 7)]:
                    sum_cell = sum_cell[0]
                    report['Detail']['L{}'.format(sum_cell.row)].value = '=$J${}-J{}'.format(gcell.row + 3, sum_cell.row)
                for sum_cell in report['Detail']['O{}:O{}'.format(gcell.row + 4, gcell.row + 7)]:
                    sum_cell = sum_cell[0]
                    report['Detail']['O{}'.format(sum_cell.row)].value = '=SUM(H{0}+L{0})'.format(sum_cell.row)
                break
        print('...Saving report...')
        report.save(bofp)
        print("\nReport saved at '{}'.".format(os.path.abspath(bofp)))
        menu = False
        break
